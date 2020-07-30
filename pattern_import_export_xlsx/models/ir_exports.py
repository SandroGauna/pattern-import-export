# Copyright 2020 Akretion France (http://www.akretion.com)
# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl).
# pylint: disable=missing-manifest-dependency
import base64
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation

from odoo import api, fields, models


class IrExports(models.Model):
    _inherit = "ir.exports"

    export_format = fields.Selection(selection_add=[("xlsx", "Excel")])

    @api.multi
    def _create_xlsx_file(self, records):
        self.ensure_one()
        book = openpyxl.Workbook()
        main_sheet = self._build_main_sheet_structure(book)
        self._populate_main_sheet_rows(main_sheet, records)
        tab_data = self.export_fields._get_tab_data()
        self._create_tabs(book, tab_data)
        main_sheet_length = len(records.ids) + 1
        self._create_validators(main_sheet, main_sheet_length, tab_data)
        book.close()
        xlsx_file = BytesIO()
        book.save(xlsx_file)
        return xlsx_file

    def _build_main_sheet_structure(self, book):
        """
        Write main sheet header and other style details
        """
        main_sheet = book["Sheet"]
        main_sheet.title = self.name
        for col, header in enumerate(self._get_header(), start=1):
            main_sheet.cell(row=1, column=col, value=header)
        return main_sheet

    def _populate_main_sheet_rows(self, main_sheet, records):
        """
        Get the actual data and write it row by row on the main sheet
        """
        for row, values in enumerate(self._get_data_to_export(records), start=2):
            for col, header in enumerate(self._get_header(), start=1):
                main_sheet.cell(row=row, column=col, value=values.get(header, ""))

    def _create_tabs(self, book, tab_data):
        """ Create additional sheets for export lines use a filter
        and write all valid choices """
        for name, headers, data, _ in tab_data:
            new_sheet = book.create_sheet(name)
            for col_number, header in enumerate(headers, start=1):
                new_sheet.cell(row=1, column=col_number, value=header)
            for row_number, row_data in enumerate(data, start=2):
                for col_number, cell_data in enumerate(row_data, start=1):
                    new_sheet.cell(row=row_number, column=col_number, value=cell_data)

    def _create_validators(self, main_sheet, main_sheet_length, tab_data):
        """ Add validators: source permitted records from tab sheets,
        apply validation to main sheet """
        for el in tab_data:
            tab_name, _, data, col_dst = el
            col_letter_dst = get_column_letter(col_dst)
            # TODO support arbitrary columns/attributes instead of
            #  only one field
            col_letter_src = get_column_letter(1)
            range_src = "${}$2:${}${}".format(
                col_letter_src, col_letter_src, str(1 + len(data))
            )
            formula_range_src = "=" + quote_sheetname(tab_name) + "!" + range_src
            validation = DataValidation(type="list", formula1=formula_range_src)
            range_dst = "${}$2:${}${}".format(
                col_letter_dst, col_letter_dst, str(main_sheet_length)
            )
            validation.add(range_dst)
            main_sheet.add_data_validation(validation)

    @api.multi
    def _export_with_record_xlsx(self, records):
        """
        Export given recordset
        @param records: recordset
        @return: string
        """
        self.ensure_one()
        excel_file = self._create_xlsx_file(records)
        return excel_file.getvalue()

    # Import part

    def _read_xlsx_file(self, datafile):
        workbook = openpyxl.load_workbook(base64.b64decode(BytesIO(datafile).read()))
        return workbook[workbook.sheetnames[0]]

    @api.multi
    def _read_import_data_xlsx(self, datafile):
        worksheet = self._read_xlsx_file(datafile)
        headers = []
        for col in range(worksheet.max_column):  # max_column is 1-based
            headers.append(worksheet.cell_value(1, col))
        for row in range(worksheet.max_row + 1):  # max_row is 1-based
            elm = {}
            for col in range(worksheet.max_column):
                elm[headers[col]] = worksheet.cell_value(row, col)
            yield elm

    def _process_load_result(self, load_result, patterned_import):
        info, status, warnings, errors = super()._process_load_result(
            self, load_result, patterned_import
        )
        if self.export_format == "xlsx":
            patterned_import.add_errors_warnings(errors, warnings)
        return info, status, warnings, errors
